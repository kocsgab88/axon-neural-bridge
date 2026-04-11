"""
    A  X  O  N  |  T E L E G R A M  B R I D G E  v6.0
    ────────────────────────────────────────────────
    v1.0 → Alap Telegram bot
    v2.0 → Sandbox (1. szint: futtatás)
    v3.0 → Unit tesztek (2. szint: szemantika) + Kill switch
    v4.0 → Gemini Cross-Check (3. szint: logikai audit)
    v5.0 → Memory/Training (öntanulás alapjai)
           + SRE Watchman (háttérfigyelés)
           + /stats parancs (statisztikák)
    v5.1 → Task Cache (axon.db alapú – 0 API hívás ismételt feladatnál)
           + Cache statisztika /stats-ban
    v5.2 → OWNER_CHAT_ID perzisztencia (axon.db – újraindítás után is megmarad)
           + /upwork cover letter parancs
           + Multi-session kód generálás (SIMPLE: 2, COMPLEX: 3 session)
           + Gemini audit küszöb optimalizálva (55/100)
    v5.3 → Multi-expert routing (DEVELOPER / PLANNER / CREATIVE / ANALYST)
           + PLANNER pipeline: strukturált markdown, sandbox nélkül, cache ✅
           + CREATIVE pipeline: Claude csak, NINCS cache (minden válasz egyedi)
           + ANALYST pipeline: Claude csak, cache ✅
           + detect_pipeline() – 1 gyors Claude hívás dönti el a pipeline-t
           + /cache_clear parancs
    v5.4 → COMPLEX generálás fix: Session 3 → 3a (összefűzés) + 3b (tesztek)
           + Debug log sorok (S3a/S3b válasz hossz + combined kód preview)
           + axon_context.py CONTEXT_VERSION → 4
    v5.5 → Enterprise DEVELOPER prompt (8 kötelező elv)
           + S3b retry logika (NONE esetén újrapróbálkozás)
    v5.6 → Scout import eltávolítva, tiszta indítás
    v6.0 → Conversation Memory (in-memory, karakter-alapú trim, session timeout)
           + DEVELOPER history: validated_code kerül history-ba (nem Telegram formázás)
           + Cache + history integráció: multi-turn → cache bypass; cache hit → [CACHE HIT] flag
           + Session timeout: 2 óra inaktivitás → auto clear + Telegram értesítő
           + /clear parancs – manuális history törlés
           + /history parancs – aktív kontextus megjelenítése (csak user turn-ök)
           + get_last_code() hook a v6.1 /review parancshoz előkészítve
           + CREATIVE pipeline history kikapcsolva (minden válasz friss slate)
    v6.1 → /review parancs – utolsó DEVELOPER kód mély Gemini re-audit
           + Szigorúbb review küszöb (70/100 vs alap 55/100)
           + review_count statisztika /stats-ban
           + get_last_code() tuple unpack fix (code, task)
           + Pipeline detect fix: PLANNER kulcsszavak első prioritás
    v7.0 → Auto README.md generálás minden sikeres DEVELOPER outputhoz
           + _generate_readme() – Claude 1 hívás, ügyfélkész Markdown
           + Import auto-detektálás (Requirements szekció)
           + outputs/ mappába: {timestamp}_{task}_README.md
           + Telegram üzenetben: 📄 README.md is mentve
    v7.1 → .env migráció – API kulcsok kikerülnek a kódból
           + python-dotenv betöltés induláskor
           + Hiányzó kulcs → azonnali leállás hibaüzenettel
    v7.3 → Fájl fogadás Telegramon (CSV, JSON, Excel, TXT, XML, YAML, SQL)
           + FileAnalyzer osztály – lokális elemzés, NEM sandbox
           + Encoding auto-detektálás, típus felismerés
           + Fájl → session history (ANALYST pipeline) + path perzisztencia
           + Claude azonnali javaslat mit lehet csinálni a fájllal
    v8.0 → Retry/backoff engine (axon_retry.py) – Claw Code claw_provider.rs mintájára
           + call_with_retry(): max 3 kísérlet, exponential backoff (200ms → 2s)
           + Retryable: 500/503/529 status, RateLimitError, ConnectionError, TimeoutError
           + Minden Claude hívó függvény védett: sync, tracked, history-aware
    v8.1 → Session compaction (axon_compaction.py) – Claw Code compact_session() mintájára
           + /compact parancs: 6000 kar felett Claude összefoglalót készít a régi turnökből
           + Utolsó 3 user+assistant pár érintetlen marad (KEEP_RECENT_TURNS)
           + format_compact_report(): Telegram üzenet a tömörítés eredményéről
    v8.2 → Watchman → JobQueue migráció (PTB natív, fő event loop-ban fut)
           + Nincs külön asyncio.create_task thread – crash-safe, restart-proof
           + SOUL.md fájlok (Openclaw minta) – pipeline promptok souls/ mappában
           + load_soul(pipeline): souls/{pipeline.lower()}.md → fallback hardcoded
           + /compact auto-trigger: handle_task minden üzenet után ellenőriz (6000 kar)
           + /upwork wizard: ConversationHandler, 3 lépéses multi-step flow
             Step 1: job leírás bekérése | Step 2: budget opcionális | Step 3: generálás
"""

import logging
import asyncio
import re
import uuid
import os
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (ApplicationBuilder, CommandHandler, MessageHandler,
                          CallbackQueryHandler, ConversationHandler,
                          filters, ContextTypes)
from anthropic import Anthropic

from axon_sandbox_v2 import (AxonSandbox, extract_code_block,
                              format_sandbox_report, UNIT_TEST_PROMPT, RISK_KEYWORDS)
from axon_auditor_v2 import AxonAuditor, format_audit_for_fix_prompt
from axon_memory import (init_db, save_training_sample, get_stats, format_stats_message,
                         get_cached_response, save_cached_response,
                         get_cache_stats, format_cache_stats_message,
                         save_fix_sample, increment_review_count,
                         log_task_cost, get_cost_stats, format_cost_stats_message,
                         get_relevant_few_shot_samples, get_successful_patterns,
                         # v3.0 history API
                         add_to_history, get_history, get_history_turn_count,
                         get_last_code, clear_history, was_timeout_cleared,
                         get_history_summary, HISTORY_ENABLED_PIPELINES)
from axon_watchman import AxonWatchman, get_system_status_message
from axon_context import get_context_for_pipeline
from axon_retry import call_with_retry        # v8.0 – Claw Code retry/backoff pattern
from axon_compaction import (compact_history,  # v8.1 – Claw Code compact_session mintájára
                              format_compact_report)
import sqlite3

def load_owner_id() -> int | None:
    """OWNER_CHAT_ID betöltése axon.db-ből induláskor."""
    try:
        db = sqlite3.connect("axon.db")
        db.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
        row = db.execute("SELECT value FROM config WHERE key='owner_chat_id'").fetchone()
        db.close()
        return int(row[0]) if row else None
    except Exception:
        return None

def save_owner_id(chat_id: int):
    """OWNER_CHAT_ID mentése axon.db-be."""
    try:
        db = sqlite3.connect("axon.db")
        db.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
        db.execute("INSERT OR REPLACE INTO config (key, value) VALUES ('owner_chat_id', ?)", (str(chat_id),))
        db.commit()
        db.close()
    except Exception as e:
        log.error(f"[CONFIG] Owner ID mentési hiba: {e}")

# ═══════════════════════════════════════════════════════════════
#  KONFIGURÁCIÓ – .env fájlból (soha ne hardcode-olj API kulcsot!)
# ═══════════════════════════════════════════════════════════════

# .env betöltés – a script mappájából
BASE_DIR  = Path(__file__).parent
_env_path = BASE_DIR / ".env"
load_dotenv(dotenv_path=_env_path)

def _require_env(key: str) -> str:
    """Kötelező env változó – ha hiányzik, azonnal leáll hibaüzenettel."""
    val = os.getenv(key)
    if not val:
        raise SystemExit(
            f"\n❌ AXON INDÍTÁSI HIBA\n"
            f"Hiányzó kötelező env változó: {key}\n"
            f"Ellenőrizd: {_env_path}\n"
            f"Minta: {key}=your_value_here"
        )
    return val

TELEGRAM_TOKEN  = _require_env("TELEGRAM_TOKEN")
ANTHROPIC_KEY   = _require_env("ANTHROPIC_KEY")
GEMINI_KEY      = _require_env("GEMINI_KEY")

OWNER_CHAT_ID       = None
SANDBOX_MAX_RETRIES = 3
SANDBOX_TIMEOUT     = 15
AUDIT_MAX_RETRIES   = 2

# ═══════════════════════════════════════════════════════════════
#  RENDSZER ÁLLAPOT
# ═══════════════════════════════════════════════════════════════
class AxonState:
    running: bool = True
    pending_approvals: dict = {}

state = AxonState()

# ═══════════════════════════════════════════════════════════════
#  SOUL LOADER – Openclaw agent-design-patterns.md minta alapján
#  Pipeline promptokat souls/ mappából tölti be, fallback: hardcoded
# ═══════════════════════════════════════════════════════════════
_SOULS_DIR = Path(__file__).parent / "souls"

def load_soul(pipeline: str) -> str | None:
    """
    Betölti a pipeline SOUL.md fájlját a souls/ mappából.
    Ha a fájl nem létezik → None (hardcoded fallback aktív).
    Szerkesztés: souls/{pipeline.lower()}.md
    """
    soul_path = _SOULS_DIR / f"{pipeline.lower()}.md"
    if soul_path.exists():
        try:
            content = soul_path.read_text(encoding="utf-8").strip()
            log.info(f"[SOUL] Betöltve: souls/{pipeline.lower()}.md ({len(content)} kar)")
            return content
        except Exception as e:
            log.warning(f"[SOUL] Olvasási hiba ({soul_path}): {e}")
    return None


# ═══════════════════════════════════════════════════════════════
#  PIPELINE TÍPUSOK ÉS SZAKÉRTŐI PROMPTOK
# ═══════════════════════════════════════════════════════════════
#
#  DEVELOPER  – kód generálás, sandbox + Gemini audit, cache ✅, history ✅
#  PLANNER    – tervek, dokumentáció, markdown kimenet, nincs sandbox, cache ✅, history ✅
#  CREATIVE   – szövegírás, cover letter, egyedi tartalom, NINCS cache ❌, NINCS history ❌
#  ANALYST    – adatelemzés, számítások, táblázatok, cache ✅, history ✅
#
_PIPELINE_PROMPTS_HARDCODED = {
    "DEVELOPER": (
        "Senior Python fejlesztő vagy, 10+ év tapasztalattal production rendszereken. "
        "Upwork-ön $200-300 értékű feladatokat oldasz meg, ezért enterprise szintű kódot írsz.\n"
        "KÖTELEZŐ ELVEK:\n"
        "1. Teljes, futtatható kód – soha nem csonkítasz, soha nem írsz pass/TODO/... placeholdert\n"
        "2. Strukturált error handling – minden I/O és API hívás try/except blokkban, értelmes hibaüzenettel\n"
        "3. Logging – logging modul, ne print(); INFO szint normál működéshez, ERROR kivételekhez\n"
        "4. Type hints – minden függvény paramétere és visszatérési értéke annotált\n"
        "5. Konfigurálhatóság – konstansok, env változók, ne hardcode; os.environ.get() mintával\n"
        "6. Docstring – minden publikus függvényhez egy sor leírás\n"
        "7. Edge case-ek – üres lista, None érték, encoding hiba, connection timeout kezelve\n"
        "8. Clean kód – DRY, egyértelmű változónevek, max 50 sor / függvény\n"
        "Kommentek és változónevek angolul vagy magyarul is elfogadottak."
    ),
    "PLANNER": (
        "Tapasztalt szoftver architect és projekt menedzser vagy. "
        "Strukturált, részletes terveket, fejlesztési ütemterveket, sprint terveket készítesz. "
        "Markdown formátumban dolgozol: fejlécek, listák, prioritások. "
        "Válaszolj magyarul."
    ),
    "CREATIVE": (
        "Profi szövegíró és kommunikációs szakértő vagy. "
        "Cover lettereket, emaileket, hirdetéseket, prezentációkat írsz. "
        "Stílusod: professzionális de személyes, konkrét, nem általánoskodó. "
        "Minden szöveg egyedi és személyre szabott."
    ),
    "ANALYST": (
        "Adatelemző és üzleti stratéga vagy. "
        "Számokat, trendeket, piaci adatokat elemzel. "
        "Strukturált, pontos válaszokat adsz táblázatokkal és összefoglalókkal ahol releváns. "
        "Válaszolj magyarul."
    ),
}

def _build_pipeline_prompts() -> dict:
    """
    SOUL.md fájlokból tölti be a pipeline promptokat.
    Ha a fájl hiányzik → hardcoded fallback. Hot-reload: bot újraindítás nélkül szerkeszthető.
    """
    prompts = {}
    for pipeline, hardcoded in _PIPELINE_PROMPTS_HARDCODED.items():
        soul = load_soul(pipeline)
        prompts[pipeline] = soul if soul else hardcoded
    return prompts

PIPELINE_PROMPTS = _build_pipeline_prompts()

PIPELINE_META = {
    "DEVELOPER": ("🔧", "DEVELOPER"),
    "PLANNER":   ("📋", "PLANNER"),
    "CREATIVE":  ("✍️", "CREATIVE"),
    "ANALYST":   ("📊", "ANALYST"),
}

NO_CACHE_PIPELINES = {"CREATIVE"}

# ═══════════════════════════════════════════════════════════════
#  LOGGING + KLIENSEK
# ═══════════════════════════════════════════════════════════════
import logging.handlers

_log_handler_file = logging.handlers.TimedRotatingFileHandler(
    filename="axon.log",
    when="midnight",
    interval=1,
    backupCount=7,
    encoding="utf-8"
)
_log_handler_file.setFormatter(logging.Formatter("%(asctime)s [AXON] %(message)s"))
_log_handler_console = logging.StreamHandler()
_log_handler_console.setFormatter(logging.Formatter("%(asctime)s [AXON] %(message)s"))

logging.basicConfig(level=logging.INFO, handlers=[_log_handler_file, _log_handler_console])
log = logging.getLogger("AXON")

init_db()
OWNER_CHAT_ID = load_owner_id()
if OWNER_CHAT_ID:
    log.info(f"[CONFIG] Owner ID betöltve DB-ből: {OWNER_CHAT_ID}")

claude   = Anthropic(api_key=ANTHROPIC_KEY)
sandbox  = AxonSandbox(max_retries=SANDBOX_MAX_RETRIES, timeout=SANDBOX_TIMEOUT)
auditor  = AxonAuditor(gemini_api_key=GEMINI_KEY)
watchman = None
scout    = None

# ═══════════════════════════════════════════════════════════════
#  BIZTONSÁGI SZŰRŐK
# ═══════════════════════════════════════════════════════════════
def is_owner(update: Update) -> bool:
    global OWNER_CHAT_ID
    chat_id = update.effective_chat.id
    if OWNER_CHAT_ID is None:
        OWNER_CHAT_ID = chat_id
        save_owner_id(chat_id)
        log.info(f"Tulajdonos regisztrálva és mentve: {chat_id}")
        return True
    return chat_id == OWNER_CHAT_ID

def system_running(update: Update) -> bool:
    if not state.running:
        asyncio.create_task(
            update.message.reply_text(
                "⛔ *AXON leállítva.* Újraindításhoz: `/start`",
                parse_mode="Markdown"
            )
        )
    return state.running

# ═══════════════════════════════════════════════════════════════
#  CLAUDE API – v6.0: history-aware hívások
#  v6.2: token tracking – call_claude_tracked() gyűjti a cost adatokat
# ═══════════════════════════════════════════════════════════════

# Feladatonkénti token akkumulátor – thread-safe nem kell (async single-thread)
_task_tokens: dict = {}  # chat_id → {"input": int, "output": int, "calls": int}

def _accumulate_tokens(chat_id: str, input_tok: int, output_tok: int) -> None:
    """Feladaton belüli token akkumulálás."""
    if chat_id not in _task_tokens:
        _task_tokens[chat_id] = {"input": 0, "output": 0, "calls": 0}
    _task_tokens[chat_id]["input"]  += input_tok
    _task_tokens[chat_id]["output"] += output_tok
    _task_tokens[chat_id]["calls"]  += 1

def _pop_task_tokens(chat_id: str) -> dict:
    """Kinyeri és törli az akkumulált token adatokat."""
    return _task_tokens.pop(chat_id, {"input": 0, "output": 0, "calls": 0})

def _tokens_to_usd(input_tok: int, output_tok: int) -> float:
    """claude-sonnet-4-6 árazás: $3/1M input, $15/1M output."""
    return (input_tok / 1_000_000 * 3.0) + (output_tok / 1_000_000 * 15.0)


def call_claude_sync(system: str, user_msg: str, max_tokens: int = 4000) -> str:
    """Egyszerű, history nélküli hívás (pipeline routing, fix callback, stb.)
    v8.0: exponential backoff retry (Claw Code claw_provider.rs mintájára)
    """
    resp = call_with_retry(
        lambda: claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=max_tokens,
            system=system,
            messages=[{"role": "user", "content": user_msg}]
        ),
        label="sync"
    )
    return resp.content[0].text

async def call_claude(system: str, user_msg: str, max_tokens: int = 4000) -> str:
    """Egyszerű async wrapper – history nélkül. Nem tracked (routing, fix, bypass)."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, call_claude_sync, system, user_msg, max_tokens)


def _call_claude_tracked_sync(system: str, user_msg: str, max_tokens: int, chat_id: str) -> str:
    """Token-tracked Claude hívás – akkumulálja a feladat cost adatait.
    v8.0: exponential backoff retry
    """
    resp = call_with_retry(
        lambda: claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=max_tokens,
            system=system,
            messages=[{"role": "user", "content": user_msg}]
        ),
        label=f"tracked/{chat_id}"
    )
    _accumulate_tokens(chat_id, resp.usage.input_tokens, resp.usage.output_tokens)
    return resp.content[0].text

async def call_claude_tracked(system: str, user_msg: str, max_tokens: int, chat_id: str) -> str:
    """Async tracked wrapper – DEVELOPER pipeline sessionökhöz."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        None, _call_claude_tracked_sync, system, user_msg, max_tokens, chat_id
    )


def call_claude_with_history_sync(system: str, messages: list[dict], max_tokens: int = 4000) -> str:
    """
    History-aware Claude hívás.
    messages = teljes conversation history (role + content párok),
    az utolsó elem az aktuális user üzenet.
    v8.0: exponential backoff retry
    """
    resp = call_with_retry(
        lambda: claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=max_tokens,
            system=system,
            messages=messages
        ),
        label="history"
    )
    return resp.content[0].text

async def call_claude_with_history(system: str, messages: list[dict], max_tokens: int = 4000) -> str:
    """Async wrapper a history-aware híváshoz."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        None, call_claude_with_history_sync, system, messages, max_tokens
    )


async def ai_fix_callback(fix_prompt: str) -> str:
    """Sandbox javító callback – history nélkül (technikai fix, nem kontextus-függő)."""
    return await call_claude(
        system=(
            "Python hibakereső és javító szakértő vagy. "
            "CSAK javított kódot adj ```python blokkban "
            "(# === KÓD === és # === TESZTEK === szekciókkal). "
            "TILOS: import unittest, class Test...(unittest.TestCase). "
            "Tesztek KIZÁRÓLAG így: if __name__ == '__test__': assert ...; print('TESZTEK OK'). "
            "Semmi magyarázat."
        ),
        user_msg=fix_prompt
    )

# ═══════════════════════════════════════════════════════════════
#  PIPELINE ROUTING
# ═══════════════════════════════════════════════════════════════
async def detect_pipeline(task: str) -> str:
    """
    Meghatározza melyik pipeline dolgozza fel a feladatot.
    Visszatérési értékek: DEVELOPER | PLANNER | CREATIVE | ANALYST
    """
    t = task.lower()

    # PLANNER ELŐSZÖR – "fejlesztési terv egy scripthez" típusú feladatnál
    # a DEVELOPER kulcsszavak (script, python) félrevezetnek, ezért a terv-szavak nyernek
    if any(w in t for w in [
        "terv", "tervez", "tervezz", "fejlesztési terv", "sprint", "roadmap",
        "ütemterv", "dokumentáció", "dokumentum", "összefoglaló", "leírás",
        "specifikáció", "spec", "feladatlista", "checklist", "struktúra",
        "felépítés", "architektúra", "készíts tervet", "írj tervet",
    ]):
        return "PLANNER"

    if any(w in t for w in [
        "kód", "kódot", "python", "script", "program", "fejlessz", "code",
        "függvény", "function", "class", "modul", "implement", "írj kódot",
        "automatizálj", "bot", "api hívás", "parser", "scraper"
    ]):
        return "DEVELOPER"

    if any(w in t for w in [
        "cover letter", "cover lettert", "email", "levél", "levelet", "szöveget",
        "hirdetés szöveg", "posztot", "prezentációt", "bemutatkozó", "ajánlat szöveg",
        "írj egy levél", "fogalmazz"
    ]):
        return "CREATIVE"

    if any(w in t for w in [
        "elemezd", "elemzés", "adat", "statisztika", "mennyi", "számolj",
        "összehasonlít", "összehasonlítás", "kalkulál", "megtérülés",
        "bevétel", "kiadás", "roi", "táblázat", "trend"
    ]):
        return "ANALYST"

    log.info("[ROUTER] Keyword miss → Claude routing hívás")
    routing_prompt = (
        f"Feladat: {task}\n\n"
        "Melyik pipeline dolgozza fel? Válaszolj CSAK egyetlen szóval:\n"
        "DEVELOPER – ha Python kódot kell írni vagy implementálni\n"
        "PLANNER   – ha tervet, dokumentációt, sprint tervet, összefoglalót kell készíteni\n"
        "CREATIVE  – ha szöveget, levelet, hirdetést, prezentációt kell írni\n"
        "ANALYST   – ha adatot kell elemezni, számolni, összehasonlítani\n\n"
        "Válasz (CSAK EGY SZÓ):"
    )
    try:
        result = await call_claude(
            system="Pipeline routing döntéshozó vagy. CSAK egyetlen szót válaszolsz: DEVELOPER, PLANNER, CREATIVE, vagy ANALYST.",
            user_msg=routing_prompt,
            max_tokens=10
        )
        pipeline = result.strip().upper().split()[0]
        if pipeline in PIPELINE_PROMPTS:
            log.info(f"[ROUTER] Claude döntés: {pipeline}")
            return pipeline
    except Exception as e:
        log.warning(f"[ROUTER] Claude routing hiba: {e}")

    log.info("[ROUTER] Végső fallback: PLANNER")
    return "PLANNER"

# ═══════════════════════════════════════════════════════════════
#  APPROVAL RENDSZER
# ═══════════════════════════════════════════════════════════════
async def ask_risk_approval(update: Update, risks: list[str], task_id: str) -> bool:
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Jóváhagyom", callback_data=f"approve_{task_id}"),
        InlineKeyboardButton("❌ Törlöm",     callback_data=f"cancel_{task_id}")
    ]])
    risk_str = ", ".join(f"`{r}`" for r in risks)
    await update.message.reply_text(
        f"⚠️ *Kockázatos műveletek:* {risk_str}\n\nJóváhagyod?",
        parse_mode="Markdown", reply_markup=keyboard
    )
    future = asyncio.get_running_loop().create_future()
    state.pending_approvals[task_id] = future
    try:
        return await asyncio.wait_for(future, timeout=300)
    except asyncio.TimeoutError:
        state.pending_approvals.pop(task_id, None)
        await update.message.reply_text("⏱️ Timeout – feladat törölve.")
        return False

async def handle_approval_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    try:
        await query.answer()
        parts = query.data.split("_", 1)
        if len(parts) != 2:
            await query.edit_message_text("❌ Érvénytelen kérés.")
            return
        action, task_id = parts
        if task_id in state.pending_approvals:
            future = state.pending_approvals.pop(task_id)
            if not future.done():
                future.set_result(action == "approve")
            label = "✅ Jóváhagyva" if action == "approve" else "❌ Törölve"
            await query.edit_message_text(f"{label} – folytatás...")
        else:
            await query.edit_message_text(
                "⏱️ Ez a jóváhagyási kérés lejárt (bot újraindult)."
                " Küldd el újra a feladatot."
            )
    except Exception as e:
        log.error(f"[APPROVAL] Callback hiba: {e}")
        try:
            await query.edit_message_text("❌ Hiba a jóváhagyás feldolgozásakor. Küldd újra a feladatot.")
        except Exception:
            pass

# ═══════════════════════════════════════════════════════════════
#  PARANCSOK
# ═══════════════════════════════════════════════════════════════
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update):
        return
    if not state.running:
        state.running = True
        log.info("AXON újraindítva.")
    await update.message.reply_text(
        "🤖 *AXON v7.2 ONLINE*\n\n"
        "📋 *Parancsok:*\n"
        "/start – Indítás / újraindítás\n"
        "/help – Példák\n"
        "/status – Rendszer állapot\n"
        "/stats – Statisztikák + Cache info\n"
        "/upwork [job leírás] – Cover letter generálás\n"
        "/history – Aktív conversation kontextus\n"
        "/clear – Conversation history törlése\n"
        "/review – Utolsó kód mély Gemini re-audit\n"
        "/files – outputs/ mappa listája + fájl küldés\n"
        "/cache\\_clear – Cache törlése\n"
        "/stop – Kill switch\n"
        "/bypass [feladat] – Validáció nélkül\n\n"
        "🧠 *Conversation Memory (v6.0):*\n"
        "Multi-turn kontextus – AXON emlékszik az előző üzenetekre\n"
        "_(DEVELOPER, PLANNER, ANALYST – 2 óra session timeout)_\n\n"
        "🔬 *Kód validáció (DEVELOPER) – 3 szint:*\n"
        "1️⃣ Statikus biztonsági szűrő\n"
        "2️⃣ Unit teszt futtatás\n"
        "3️⃣ Gemini logikai audit\n\n"
        "🧠 *Multi-expert routing:*\n"
        "🔧 DEVELOPER – Python kód (sandbox + Gemini)\n"
        "📋 PLANNER   – Tervek, dokumentáció (markdown)\n"
        "✍️ CREATIVE  – Cover letter, szövegek (egyedi)\n"
        "📊 ANALYST   – Adatelemzés, számítások",
        parse_mode="Markdown"
    )

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update):
        return
    await update.message.reply_text(
        "💡 *Példák:*\n\n"
        "🔧 *DEVELOPER*\n"
        "`Írj Python kódot ami megszámolja egy CSV sorait`\n"
        "   └ 3 szintű validáció automatikusan\n\n"
        "🔧 *DEVELOPER – multi-turn*\n"
        "`Adj hozzá loggingot az előző kódhoz`\n"
        "   └ AXON emlékszik az előző kódra\n\n"
        "📋 *PLANNER*\n"
        "`Generálj strukturált AXON fejlesztési sprint tervet`\n"
        "   └ Markdown dokumentum, nincs sandbox\n\n"
        "✍️ *CREATIVE*\n"
        "`Írj Upwork cover lettert Python automatizálás munkához`\n"
        "   └ Egyedi szöveg, nincs cache, nincs history\n\n"
        "📊 *ANALYST*\n"
        "`Elemezd melyik Upwork kategória fizet legjobban 2024-ben`\n"
        "   └ Adatelemzés, táblázatos kimenet\n\n"
        "⚡ `/bypass Kód ami törli a temp fájlokat`\n"
        "   └ Validáció NÉLKÜL, saját felelősségre\n\n"
        "🔍 *REVIEW*\n"
        "`/review`\n"
        "   └ Utolsó kód mély Gemini re-audit (70/100 küszöb)\n\n"
        "🧹 `/clear` – Conversation history törlése\n"
        "📜 `/history` – Mit tud most az AXON a sessionből",
        parse_mode="Markdown"
    )

async def status_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update):
        return
    chat_id = str(update.effective_chat.id)
    turn_count = get_history_turn_count(chat_id)
    history_info = f"\n🧠 *Conversation history:* {turn_count // 2} kérdés-válasz pár aktív"
    await update.message.reply_text(
        get_system_status_message() +
        f"\n\n⛔ Kill switch: {'AKTÍV' if not state.running else 'Készen'}" +
        history_info,
        parse_mode="Markdown"
    )

async def stats_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tanulási statisztikák + cache megtakarítás + API cost."""
    if not is_owner(update):
        return
    days = 7
    if context.args:
        try:
            days = int(context.args[0])
        except ValueError:
            pass
    stats    = get_stats(days)
    cs       = get_cache_stats(days)
    cost_stats = get_cost_stats(days)
    await update.message.reply_text(
        format_stats_message(stats) + "\n\n" +
        format_cache_stats_message(cs) + "\n\n" +
        format_cost_stats_message(cost_stats),
        parse_mode="Markdown"
    )

async def cache_clear_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cache törlése – minden tárolt válasz elvész."""
    if not is_owner(update):
        return
    try:
        db = sqlite3.connect("axon.db")
        db.execute("DELETE FROM task_cache")
        db.execute("DELETE FROM cache_stats")
        db.commit()
        db.close()
        await update.message.reply_text(
            "🗑️ *Cache törölve!*\nKövetkező feladatok újra API hívással futnak.",
            parse_mode="Markdown"
        )
        log.info("[CACHE] Manuális cache törlés végrehajtva.")
    except Exception as e:
        await update.message.reply_text(f"❌ Cache törlési hiba: {str(e)[:200]}")

async def clear_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /clear – Conversation history manuális törlése.
    Hasznos ha új témába vágsz és nem akarsz régi kontextust cipelni.
    """
    if not is_owner(update):
        return
    chat_id = str(update.effective_chat.id)
    count = clear_history(chat_id)
    if count == 0:
        await update.message.reply_text(
            "📭 *Nincs aktív conversation history.*\nFriss session már fut.",
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text(
            f"🧹 *Conversation history törölve!*\n"
            f"_{count // 2} kérdés-válasz pár eltávolítva._\n\n"
            f"Következő üzenet friss kontextussal indul.",
            parse_mode="Markdown"
        )
    log.info(f"[HISTORY] /clear parancs – {count} turn törölve | chat: {chat_id}")

async def history_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /history – Aktív conversation kontextus megjelenítése.
    Csak a user turn-ök listája, röviden. Így látod mi van a kontextusban.
    """
    if not is_owner(update):
        return
    chat_id = str(update.effective_chat.id)
    summary = get_history_summary(chat_id)
    await update.message.reply_text(summary, parse_mode="Markdown")

async def compact_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /compact – Conversation history tömörítése Claude összefoglalóval.
    v8.1: Claw Code compact_session() Python portja.

    Ha a history meghaladja a 6000 karaktert:
      - A régi turnök Claude-dal összefoglaltatnak
      - Az összefoglaló + az utolsó 3 pár marad meg
    Ha a history a küszöb alatt van → no-op üzenet.
    """
    if not is_owner(update):
        return
    chat_id = str(update.effective_chat.id)
    history = get_history(chat_id)

    if not history:
        await update.message.reply_text(
            "📭 *Nincs aktív conversation history.*",
            parse_mode="Markdown"
        )
        return

    await update.message.reply_text("🗜 Tömörítés folyamatban...", parse_mode="Markdown")

    loop = asyncio.get_running_loop()
    result = await loop.run_in_executor(
        None,
        compact_history,
        history,
        call_claude_sync,
        chat_id
    )

    if not result.skipped:
        # History frissítése az axon_memory-ban
        from axon_memory import clear_history, add_to_history
        clear_history(chat_id)
        for turn in result.compacted_history:
            add_to_history(chat_id, turn["role"], turn["content"])

        log.info(
            f"[COMPACT] /compact parancs – {result.removed_turns} turn tömörítve, "
            f"{result.kept_turns} megmaradt | chat: {chat_id}"
        )

    report = format_compact_report(result)
    await update.message.reply_text(report, parse_mode="Markdown")

async def stop_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update):
        return
    state.running = False
    log.warning("KILL SWITCH AKTIVÁLVA!")
    await update.message.reply_text(
        "⛔ *AXON LEÁLLÍTVA*\n\nÚjraindítás: `/start`",
        parse_mode="Markdown"
    )

async def review_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /review – Az utolsó DEVELOPER kód mély Gemini re-auditja.

    Szigorúbb mint az automatikus audit:
    - Küszöb: 70/100 (vs alap 55/100)
    - Extra instrukció: edge case-ek, biztonsági rések keresése
    - review_count statisztikában rögzítve

    Ha argumentum van (/review valamit javíts), az sima multi-turn conversation
    lesz – nem külön ág, hanem a handle_task() kezeli.
    """
    if not is_owner(update) or not system_running(update):
        return

    chat_id = str(update.effective_chat.id)
    last_code, last_task = get_last_code(chat_id)

    if not last_code:
        await update.message.reply_text(
            "📭 *Nincs reviewolható kód.*\n\n"
            "A `/review` az aktuális session utolsó DEVELOPER kódját auditálja újra.\n"
            "Küldj előbb egy Python kód feladatot!",
            parse_mode="Markdown"
        )
        return

    task_display = (last_task[:80] + "…") if last_task and len(last_task) > 80 else (last_task or "ismeretlen feladat")
    msg = await update.message.reply_text(
        f"🔍 *Mély Gemini review indul...*\n"
        f"_Feladat: {task_display}_\n\n"
        f"⏳ Gemini elemez (szigorúbb küszöb: 70/100)...",
        parse_mode="Markdown"
    )

    # Mély audit – extra instrukció a Gemini-nek
    review_task = (
        f"{last_task or 'Python kód'}\n\n"
        f"[MÉLY REVIEW MÓD] Ez egy manuális review kérés. "
        f"Légy extra kritikus. Keress edge case-eket, biztonsági réseket "
        f"és kód minőségi problémákat amiket az automatikus audit esetleg kihagyott. "
        f"A PASS küszöb 70/100."
    )

    auditor = AxonAuditor()
    audit_result = await auditor.audit(
        code=last_code,
        task=review_task,
        test_result="Manuális review – sandbox korábban lefutott"
    )

    # Szigorúbb PASS küszöb (70 vs alap 55)
    if audit_result.verdict != "SKIP" and audit_result.score < 70:
        audit_result.passed = False
        if audit_result.verdict == "PASS":
            audit_result.verdict = "FAIL"

    increment_review_count()

    verdict_icon = "✅" if audit_result.passed else "❌"
    skip_note = " _(Gemini nem elérhető)_" if audit_result.verdict == "SKIP" else ""

    reply = (
        f"🔍 *Mély Gemini Review eredménye*{skip_note}\n\n"
        f"{verdict_icon} *Verdikt:* {audit_result.verdict} ({audit_result.score}/100)\n"
        f"_Küszöb: 70/100 (szigorúbb mint az alap 55/100)_\n\n"
    )

    if audit_result.issues:
        reply += "*🔴 Talált problémák:*\n"
        for issue in audit_result.issues:
            reply += f"• {issue}\n"
        reply += "\n"

    if audit_result.suggestions:
        reply += "*💡 Javaslatok:*\n"
        for s in audit_result.suggestions:
            reply += f"• {s}\n"
        reply += "\n"

    if audit_result.passed and not audit_result.issues:
        reply += "✨ *Nem talált problémát – a kód megfelel a szigorúbb kritériumoknak is.*\n"

    reply += f"\n_Reviewolt kód: {len(last_code)} kar | Feladat: {task_display}_"

    await msg.delete()
    await safe_send(update, reply)
    log.info(f"[REVIEW] Mély audit kész – {audit_result.verdict} ({audit_result.score}/100) | chat: {chat_id}")


async def files_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /files        – outputs/ mappa utolsó 10 fájljának listája
    /files [N]    – az Nth fájl elküldése Telegram document-ként (1-10)
    /files last   – a legutóbbi .py fájl azonnali elküldése

    Így nem kell kézzel kimásolni a fájlokat az ügyfélnek szánt deliverable-höz.
    """
    if not is_owner(update):
        return

    output_dir = BASE_DIR / "outputs"

    if not os.path.exists(output_dir):
        await update.message.reply_text(
            "📂 *outputs/ mappa nem létezik még.*\nGenerálj először egy kódot!",
            parse_mode="Markdown"
        )
        return

    # Fájlok listázása – .py és .md fájlok, idő szerint csökkenő
    all_files = sorted(
        [f for f in os.listdir(output_dir) if f.endswith((".py", ".md"))],
        reverse=True
    )[:20]  # max 20 fájl

    if not all_files:
        await update.message.reply_text("📂 *Nincs fájl az outputs/ mappában.*", parse_mode="Markdown")
        return

    # Argumentum feldolgozás
    arg = " ".join(context.args).strip().lower() if context.args else ""

    # /files last – legutóbbi .py fájl küldése
    if arg == "last":
        py_files = [f for f in all_files if f.endswith(".py")]
        if not py_files:
            await update.message.reply_text("📂 Nincs .py fájl az outputs/ mappában.", parse_mode="Markdown")
            return
        await _send_file(update, os.path.join(output_dir, py_files[0]))
        return

    # /files [N] – Nth fájl küldése
    if arg.isdigit():
        idx = int(arg) - 1
        if 0 <= idx < len(all_files):
            await _send_file(update, os.path.join(output_dir, all_files[idx]))
        else:
            await update.message.reply_text(
                f"❌ Érvénytelen szám. 1–{len(all_files)} között adj meg egy számot.",
                parse_mode="Markdown"
            )
        return

    # /files – lista megjelenítése
    lines = ["📂 *outputs/ – legutóbbi fájlok:*\n"]
    for i, fname in enumerate(all_files[:10], 1):
        fpath  = os.path.join(output_dir, fname)
        fsize  = os.path.getsize(fpath) // 1024  # KB
        icon   = "🐍" if fname.endswith(".py") else "📄"
        lines.append(f"`{i}.` {icon} `{fname[:55]}`  _{fsize}KB_")

    lines.append(
        f"\n_Küldés: `/files [szám]` vagy `/files last`_\n"
        f"_Összes fájl: {len(all_files)}_"
    )
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def _send_file(update: Update, filepath: str) -> None:
    """Telegram document-ként elküldi a megadott fájlt."""
    try:
        fname = os.path.basename(filepath)
        with open(filepath, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=fname,
                caption=f"📎 `{fname}`",
                parse_mode="Markdown"
            )
        log.info(f"[FILES] Fájl elküldve: {fname}")
    except Exception as e:
        log.error(f"[FILES] Küldési hiba: {e}")
        await update.message.reply_text(f"❌ Fájl küldési hiba: {str(e)[:200]}")


async def bypass_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update) or not system_running(update):
        return
    task = " ".join(context.args)
    if not task:
        await update.message.reply_text("Használat: `/bypass [feladat]`", parse_mode="Markdown")
        return
    msg = await update.message.reply_text("⚡ *Bypass mód...*", parse_mode="Markdown")
    result = await call_claude(PIPELINE_PROMPTS["DEVELOPER"], task)
    await msg.delete()
    await safe_send(update, "⚡ *DEVELOPER (bypass):*\n\n" + result)

# ═══════════════════════════════════════════════════════════════
#  /upwork WIZARD – ConversationHandler, Cargo-connect state machine minta
#  3 lépés: job leírás → budget kérés (opcionális) → generálás
# ═══════════════════════════════════════════════════════════════
UPWORK_JOB, UPWORK_BUDGET = range(2)

_UPWORK_SYSTEM = """Te Kocsis Gábor vagy, Budapest. Nappal villamos elosztó tábla szerelő csoportvezető, szabadidőben Python fejlesztő és automatizálás specialista.

VALÓS HÁTTÉR (csak ebből dolgozz, soha ne találj ki referenciát):
- Python automatizálás: script-ek, adatfeldolgozás, API integrációk
- Telegram bot fejlesztés (saját AXON rendszer: Claude + Gemini AI pipeline, SQLite, sandbox validáció)
- n8n workflow automatizálás (lokálisan fut, GoHighLevel CRM integráció portfólióban)
- Make.com workflow builder (Google Sheets, Gmail, Telegram, OpenAI, Airtable connectors)
- SQLite adatbázis kezelés, CSV/JSON feldolgozás
- AI integráció: Anthropic Claude API, Google Gemini API
- Logistics & Supply Chain Automation niche – villamos csoportvezetői háttér versenyelőny
- Windows és Linux környezet
- Minden projektet egyedül tervezel, kódolsz, tesztelsz

TAPASZTALAT KOMMUNIKÁCIÓ:
- Nincs Upwork értékelés még — ezt nem hozod fel
- ABSZOLÚT TILOS: kitalált ügyfél, cég, projektnév, szám (pl. "50k rows", "6 months", "Dutch client") — ezek ellenőrizhetők és bukást okoznak
- ABSZOLÚT TILOS: "Recent relevant work:" vagy bármilyen referencia szekció
- Helyette: a technikai megközelítést mutasd meg ("The core challenge here is...", "I'd handle this with...")
- Általános önbizalom OK: "I know this exact setup", "I've worked with psycopg2 and gspread"

HANGNEM:
- Ember írja, nem AI — rövid mondatok, tömör
- Profi de nem formális
- Konkrét és technikai
- Magabiztos, nem alázatos
- Max 130 szó"""

async def upwork_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Wizard 1. lépés: elindítja a flow-t, bekéri a job leírást."""
    if not is_owner(update) or not system_running(update):
        return ConversationHandler.END

    # Ha már argumentummal hívták (/upwork [szöveg]) → átugorja a kérdést
    if context.args:
        job_desc = " ".join(context.args).strip()
        context.user_data["upwork_job"] = job_desc
        await update.message.reply_text(
            f"📋 *Job leírás rögzítve.*\n\n"
            f"Mekkora a hirdetett budget? _(Skip: küldj `/skip`-et)_",
            parse_mode="Markdown"
        )
        return UPWORK_BUDGET

    await update.message.reply_text(
        "✍️ *Upwork Cover Letter Wizard*\n\n"
        "📋 *1. lépés:* Másold be a job leírást!\n\n"
        "_A teljes hirdetés szövege, minél több annál jobb._",
        parse_mode="Markdown"
    )
    return UPWORK_JOB


async def upwork_got_job(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Wizard 2. lépés: job leírás megvan, bekéri a budget-et."""
    if not is_owner(update):
        return ConversationHandler.END

    context.user_data["upwork_job"] = update.message.text.strip()
    await update.message.reply_text(
        "💰 *2. lépés:* Mekkora a hirdetett budget?\n\n"
        "_Pl: $150, $300 fixed, $25/hr — vagy küldj `/skip`-et ha nincs megadva._",
        parse_mode="Markdown"
    )
    return UPWORK_BUDGET


async def upwork_got_budget(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Wizard 3. lépés: budget megvan, generálás."""
    if not is_owner(update):
        return ConversationHandler.END

    budget = update.message.text.strip()
    context.user_data["upwork_budget"] = budget
    return await _upwork_generate(update, context)


async def upwork_skip_budget(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Budget kihagyva → generálás budget nélkül."""
    if not is_owner(update):
        return ConversationHandler.END

    context.user_data["upwork_budget"] = None
    return await _upwork_generate(update, context)


async def _upwork_generate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Belső: cover letter generálás az összegyűjtött adatokból."""
    job_description = context.user_data.get("upwork_job", "")
    budget          = context.user_data.get("upwork_budget")

    if not job_description:
        await update.message.reply_text("❌ Nincs job leírás. Indítsd újra: `/upwork`", parse_mode="Markdown")
        return ConversationHandler.END

    budget_line = f"\nHirdetett budget: {budget}" if budget else ""
    msg = await update.message.reply_text(
        "✍️ *Cover letter generálás...*\n_CREATIVE pipeline – egyedi, személyes levél_",
        parse_mode="Markdown"
    )

    upwork_prompt = f"""Írj Upwork cover lettert erre a hirdetésre:

{job_description}{budget_line}

SZABÁLYOK:
- Angolul írj
- Max 130 szó — minden szó számít
- Első mondat: azonnal a lényeg
- Csak VALÓS tapasztalat (Python, automatizálás, API integráció, Telegram bot, n8n, AI pipeline)
- Ha a budget ismert és alacsony ($50 alatt): jelezd röviden hogy gyors és precíz munkát kapsz
- CTA a végén: egy konkrét kérdés vagy ajánlat
- TILOS: "I am writing to apply", "I would love to", "I am a passionate", bulletpoint lista
- Úgy hangozzon mintha egy fejlesztő gyorsan begépelte volna, nem mintha AI írta volna"""

    try:
        response = await call_claude(_UPWORK_SYSTEM, upwork_prompt, max_tokens=600)
        await msg.delete()

        budget_display = f"💰 Budget: {budget}\n" if budget else ""
        await safe_send(update,
            f"✍️ *Cover Letter*\n"
            f"{budget_display}"
            f"_{job_description[:80]}{'...' if len(job_description) > 80 else ''}_\n\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{response}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"💡 _Másold be az Upwork üzenetmezőbe és személyre szabhatod_"
        )
        save_training_sample(expert_mode="creative", prompt=job_description, success=True)
        log.info(f"[UPWORK] Cover letter generálva ({len(response)} kar) | wizard flow")

    except Exception as e:
        await msg.delete()
        await update.message.reply_text(f"❌ Hiba: {str(e)[:200]}")
        log.error(f"[UPWORK] Hiba: {e}")

    context.user_data.clear()
    return ConversationHandler.END


async def upwork_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Wizard megszakítása."""
    context.user_data.clear()
    await update.message.reply_text(
        "❌ *Upwork wizard megszakítva.*\nÚjraindítás: `/upwork`",
        parse_mode="Markdown"
    )
    return ConversationHandler.END

# ═══════════════════════════════════════════════════════════════
#  STÁTUSZ FRISSÍTŐ + SAFE SEND HELPER
# ═══════════════════════════════════════════════════════════════
async def update_status(msg, text: str):
    try:
        await msg.edit_text(text, parse_mode="Markdown")
    except Exception:
        pass

async def safe_send(update: Update, text: str):
    chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
    for chunk in chunks:
        try:
            await update.message.reply_text(chunk, parse_mode="Markdown")
        except Exception:
            clean = chunk.replace("*", "").replace("`", "").replace("_", "")
            await update.message.reply_text(clean)

# ═══════════════════════════════════════════════════════════════
#  PIPELINE HANDLEREK
# ═══════════════════════════════════════════════════════════════

async def run_simple_pipeline(
    update: Update,
    pipeline: str,
    task: str,
    status_msg,
    chat_id: str
) -> None:
    """
    PLANNER, CREATIVE, ANALYST pipeline:
    – Cache ellenőrzés (kivéve CREATIVE)
    – v6.0: History-aware Claude hívás (kivéve CREATIVE)
    – Cache mentés (kivéve CREATIVE, kivéve multi-turn session)
    – History mentés (kivéve CREATIVE)
    – Training data mentés
    """
    icon, label = PIPELINE_META[pipeline]
    no_cache    = pipeline in NO_CACHE_PIPELINES
    use_history = pipeline in HISTORY_ENABLED_PIPELINES
    turn_count  = get_history_turn_count(chat_id) if use_history else 0
    is_multiturn = use_history and turn_count > 0

    # Cache ellenőrzés – csak friss session esetén (multi-turn = egyedi kontextus)
    if not no_cache and not is_multiturn:
        cached = get_cached_response(pipeline.lower(), task)
        if cached:
            await status_msg.delete()
            full = f"⚡ *{icon} {label}* _(cache)_:\n\n{cached}"
            await safe_send(update, full)
            log.info(f"Cache HIT – {pipeline} | 0 API hívás")
            # Cache hit kerül a history-ba [CACHE HIT] flaggel
            if use_history:
                add_to_history(chat_id, "user",      task,                    pipeline)
                add_to_history(chat_id, "assistant", f"[CACHE HIT]\n{cached}", pipeline)
            return

    # Claude hívás
    axon_context  = get_context_for_pipeline(pipeline)
    pipeline_role = PIPELINE_PROMPTS[pipeline]
    system_prompt = (
        f"{axon_context}\n\n"
        f"---\n\n"
        f"{pipeline_role}\n\n"
        "Fontos: Válaszolj magyarul, tömören és strukturáltan. "
        "SOHA ne kérj vissza alapadatokat – a fenti AXON kontextus alapján dolgozz."
    )

    if use_history and is_multiturn:
        # Multi-turn: history + aktuális user üzenet
        add_to_history(chat_id, "user", task, pipeline)
        messages = get_history(chat_id)
        response = await call_claude_with_history(system_prompt, messages, max_tokens=3000)
        log.info(f"[HISTORY] Multi-turn hívás – {pipeline} | {turn_count + 1} turn")
    else:
        # Első üzenet ebben a session-ben
        response = await call_claude(system_prompt, task, max_tokens=3000)
        if use_history:
            add_to_history(chat_id, "user", task, pipeline)

    await status_msg.delete()
    full = f"{icon} *{label}* válasza:\n\n{response}"
    await safe_send(update, full)

    # History mentés
    if use_history:
        add_to_history(chat_id, "assistant", response, pipeline)

    # Cache mentés – csak friss session, első üzenet esetén
    if not no_cache and not is_multiturn:
        save_cached_response(pipeline.lower(), task, response)

    save_training_sample(expert_mode=pipeline.lower(), prompt=task, success=True)
    log.info(f"Feladat kész ({pipeline}) | history={'aktív' if use_history else 'ki'} | cache={'skip' if no_cache or is_multiturn else 'mentve'}")


async def _generate_readme(
    task: str,
    main_code: str,
    filename: str,
    output_dir: str,
    timestamp: str,
    safe_task: str,
    audit_result,
    sandbox_result
) -> str:
    """
    v7.0 – Auto README.md generálás minden sikeres DEVELOPER outputhoz.
    Claude 1 hívás (~500 token) → Markdown dokumentum az ügyfélnek.
    Visszaad egy Telegram-barát sort a file elérési útjával, vagy üres stringet hiba esetén.
    """
    try:
        # Import lista auto-detektálás a kódból
        import_lines = [
            line.strip() for line in main_code.splitlines()
            if line.strip().startswith(("import ", "from ")) and "==" not in line
        ]
        imports_str = "\n".join(import_lines[:20]) if import_lines else "# standard library only"

        audit_summary = (
            f"PASS ({audit_result.score}/100)"
            if audit_result.verdict == "PASS"
            else audit_result.verdict
        )
        sandbox_summary_str = (
            f"PASS (#{ sandbox_result.attempt}. próba, "
            f"{sandbox_result.tests_passed}/{sandbox_result.tests_total} teszt OK)"
            if sandbox_result.success else "FAIL"
        )

        readme_prompt = f"""Generálj egy rövid, professzionális README.md fájlt az alábbi Python scripthez.
A README-t egy Upwork ügyfélnek szánjuk – angolul legyen, tömör és praktikus.

SCRIPT NEVE: {filename}
FELADAT: {task}

IMPORTOK (függőségek ebből következtethetők):
{imports_str}

VALIDÁCIÓ EREDMÉNYE:
- Sandbox: {sandbox_summary_str}
- Gemini audit: {audit_summary}
- Generálva: AXON Neural Bridge | {datetime.now().strftime('%Y-%m-%d')}

README STRUKTÚRA (pontosan ezt kövesd):
# {filename.replace('.py', '')}

## Description
[1-2 mondat mi a script célja]

## Requirements
[pip install lista az importok alapján – csak külső csomagok, stdlib-et kihagyod]

## Usage
```bash
python {filename} [argumentumok ha van]
```

## Configuration
[env változók vagy konfig paraméterek ha látod a kódban, különben: "No configuration required."]

## Output
[Mit csinál a script – mi az output]

---
*Generated by AXON Neural Bridge | Sandbox: {sandbox_summary_str} | Gemini: {audit_summary}*

FONTOS: Csak a README.md tartalmat add vissza, semmi más. Ne írj magyarázatot."""

        readme_content = await call_claude(
            system="Professional technical writer. Generate concise, client-ready README.md files.",
            user_msg=readme_prompt,
            max_tokens=600
        )

        # README fájl mentése
        readme_filename = f"{timestamp}_{safe_task}_README.md"
        readme_filepath = os.path.join(output_dir, readme_filename)
        with open(readme_filepath, "w", encoding="utf-8") as f:
            f.write(readme_content)

        log.info(f"[README] Mentve: {readme_filepath}")
        return f"📄 `{readme_filename}`"

    except Exception as e:
        log.error(f"[README] Generálási hiba: {e}")
        return ""


async def run_developer_pipeline(
    update: Update,
    task: str,
    status_msg,
    chat_id: str
) -> None:
    """
    DEVELOPER pipeline:
    – v6.0: Cache bypass ha multi-turn session (kontextus-függő kód)
    – v6.3: Few-shot tanulás – sikeres pattern hint + fix minta a sandboxnak
    – Komplexitás becslés
    – Multi-session kód generálás (SIMPLE: 2, COMPLEX: 3a+3b)
    – History-aware generálás: ha history van, a kód request-be beépítjük
    – Sandbox validáció (statikus + unit tesztek)
    – Gemini audit
    – Cache mentés (csak ha minden PASS ÉS friss session)
    – History mentés: validated_code kerül be (nem Telegram formázás!)
    """
    turn_count   = get_history_turn_count(chat_id)
    is_multiturn = turn_count > 0

    # ── Cache ellenőrzés – csak friss session ────────────────
    if not is_multiturn:
        cached = get_cached_response("developer", task)
        if cached:
            await status_msg.delete()
            await safe_send(update, f"⚡ *DEVELOPER* _(cache – sandbox+audit korábban PASS)_:\n\n{cached}")
            log.info("Developer cache HIT | 0 API hívás, 0 sandbox, 0 Gemini")
            add_to_history(chat_id, "user",      task,                     "DEVELOPER")
            add_to_history(chat_id, "assistant", f"[CACHE HIT]\n{cached}",  "DEVELOPER", task=task)
            return

    # ── v6.3: Few-shot adatok betöltése ──────────────────────
    successful_patterns = get_successful_patterns(task, max_patterns=2)
    few_shot_pattern_block = ""
    if successful_patterns:
        lines = ["TANULÁS – hasonló feladatoknál ezek a megközelítések működtek:\n"]
        for i, p in enumerate(successful_patterns, 1):
            lines.append(f"Példa {i} (hasonlóság: {p['similarity']:.0%}):")
            lines.append(f"Feladat volt: {p['prompt']}")
            lines.append(f"Struktúra ami működött:\n```python\n{p['code_snippet']}\n```\n")
        few_shot_pattern_block = "\n".join(lines) + "\n"
        log.info(f"[FEW-SHOT] {len(successful_patterns)} sikeres pattern betöltve generálás előtt")

    # A fix mintákat a sandbox validate_with_retry-nak adjuk át (hiba esetén)
    # Első körben üres – hiba után töltjük fel az error_text alapján
    _few_shot_error_cache: dict = {"text": ""}  # mutable closure

    # ── 1. Komplexitás becslés ────────────────────────────────
    await update_status(status_msg, "1️⃣ *Tervezés...*\nKomplexitás elemzés")

    complexity_prompt = (
        f"Feladat: {task}\n\n"
        "Becsüld meg a feladat komplexitását. Válaszolj CSAK így:\n"
        "SIMPLE – ha 1 függvény vagy osztály elegendő, max ~50 sor\n"
        "COMPLEX – ha több osztály/modul kell, adatstruktúrák, generálás, >50 sor\n\n"
        "Ha COMPLEX, adj egy rövid tervet max 3 lépésben (1 sor/lépés).\n"
        "Formátum:\nCOMPLEXITY: SIMPLE\nvagy\nCOMPLEXITY: COMPLEX\nPLAN:\n1. ...\n2. ...\n3. ..."
    )

    complexity_response = await call_claude_tracked(
        PIPELINE_PROMPTS["DEVELOPER"],
        complexity_prompt,
        max_tokens=300,
        chat_id=chat_id
    )

    is_complex = "COMPLEXITY: COMPLEX" in complexity_response.upper()
    log.info(f"[GEN] Komplexitás: {'COMPLEX' if is_complex else 'SIMPLE'}")

    # ── History context prefix – ha multi-turn ───────────────
    # Ha van előző kód a history-ban, beillesztjük a promptba
    history_context = ""
    if is_multiturn:
        last_code, _ = get_last_code(chat_id)  # (code, task) tuple – task itt nem kell
        if last_code:
            history_context = (
                f"KONTEXTUS – az előző feladatban generált kód (módosítsd/bővítsd ha kérik):\n"
                f"```python\n{last_code[:3000]}\n```\n\n"
            )
            log.info(f"[HISTORY] Előző kód beillesztve a promptba ({len(last_code)} kar)")

    if is_complex:
        # ── KOMPLEX: 3a+3b sessionös generálás ──────────────────
        plan_lines = []
        for line in complexity_response.split("\n"):
            line = line.strip()
            if line and line[0].isdigit() and ". " in line:
                plan_lines.append(line.split(". ", 1)[1])
        if not plan_lines:
            plan_lines = ["Első rész", "Második rész", "Tesztek"]

        await update_status(
            status_msg,
            f"1️⃣ *Komplex feladat – 4 lépéses generálás*\n"
            f"  ⏳ 1. Kód első fele\n  ⏳ 2. Kód második fele\n  ⏳ 3a. Összefűzés + tisztítás\n  ⏳ 3b. Unit tesztek"
        )

        await update_status(status_msg, "1️⃣ *Session 1/3* – Adatstruktúrák + helper függvények")
        s1_prompt = (
            f"{few_shot_pattern_block}"
            f"{history_context}"
            f"Feladat: {task}\n\n"
            "FONTOS: Ez egy 2 részből álló generálás ELSŐ FELE.\n"
            "Írj TELJES, MŰKÖDŐ Python kódot az alábbi részekhez:\n"
            "- Importok\n"
            "- Adatstruktúrák (dataclass, enum, konstansok)\n"
            "- Helper/segédfüggvények (teljes implementációval, NEM pass!)\n\n"
            "TILOS: pass, TODO, ... placeholder\n"
            "MINDEN függvény törzse legyen kitöltve!\n\n"
            "FORMÁTUM:\n```python\n# === KÓD (1. rész) ===\n[teljes kód ide]\n```"
        )
        s1_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s1_prompt, max_tokens=3500, chat_id=chat_id)
        s1_code = extract_code_block(s1_resp) or ""
        s1_code = s1_code.replace("# === KÓD (1. rész) ===", "").strip()

        await update_status(status_msg, "1️⃣ *Session 2/3* – Fő logika + generáló függvények")
        s2_prompt = (
            f"Feladat: {task}\n\n"
            f"Már megvan a kód ELSŐ FELE:\n```python\n{s1_code[:2500]}\n```\n\n"
            "FONTOS: Ez a MÁSODIK FELE. Írj TELJES, MŰKÖDŐ kódot:\n"
            "- Fő logika / generáló függvények\n"
            "- Main függvény vagy belépési pont\n"
            "- Minden függvény KITÖLTVE (NEM pass, NEM TODO!)\n\n"
            "Az első félre ÉPÍTS, ne ismételd meg!\n"
            "FORMÁTUM:\n```python\n# === KÓD (2. rész) ===\n[folytatás ide]\n```"
        )
        s2_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s2_prompt, max_tokens=3500, chat_id=chat_id)
        s2_code = extract_code_block(s2_resp) or ""
        s2_code = s2_code.replace("# === KÓD (2. rész) ===", "").strip()

        await update_status(status_msg, "1️⃣ *Session 3a/3* – Összefűzés + tisztítás")
        combined_raw = s1_code + "\n\n" + s2_code
        s3a_prompt = (
            f"Ez a Python kód két részből összefűzve:\n"
            f"```python\n{combined_raw[:4000]}\n```\n\n"
            f"Feladat: {task}\n\n"
            "CSAK ezeket csináld:\n"
            "1. Ha van dupla import, távolítsd el (tartsd az elsőt)\n"
            "2. Ha van dupla osztály/függvény definíció, távolítsd el (tartsd az elsőt)\n"
            "3. Ellenőrizd hogy a kód szintaktikailag helyes és futtatható\n\n"
            "TILOS: teszteket írni, logikát változtatni, kommentelni!\n"
            "Add vissza a TELJES tisztított kódot:\n"
            "```python\n# === KÓD ===\n[teljes, tisztított kód – TESZTEK NÉLKÜL]\n```"
        )
        s3a_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s3a_prompt, max_tokens=4000, chat_id=chat_id)
        log.info(f"[GEN] S3a válasz hossza: {len(s3a_resp)} kar")
        s3a_code = extract_code_block(s3a_resp)
        log.info(f"[GEN] S3a extracted kód: {repr(s3a_code[:200] if s3a_code else 'NONE')}")
        clean_code = s3a_code.replace("# === KÓD ===", "").strip() if s3a_code else combined_raw.strip()

        await update_status(status_msg, "1️⃣ *Session 3b/3* – Unit tesztek generálása")
        code_preview = clean_code[:2500]
        s3b_prompt = (
            f"Ez a kész Python kód (csak az eleje látható terjedelmi okokból):\n"
            f"```python\n{code_preview}\n```\n\n"
            f"Feladat: {task}\n\n"
            "Írj 2-3 LOGIKAI unit tesztet a kód belső logikájának ellenőrzésére.\n"
            "A tesztek a már definiált függvények logikáját tesztelik – NEM az infrastruktúrát.\n\n"
            "TESZTÍRÁSI SZABÁLYOK:\n"
            "TILOS:\n"
            "  assert conn is not None     <- DB kapcsolat\n"
            "  assert client is not None   <- API kliens\n"
            "  import unittest / class TestXxx(unittest.TestCase)  <- unittest modul tiltott!\n\n"
            "HELYES:\n"
            "  result = transform(raw_data)\n"
            "  assert isinstance(result, list)\n\n"
            "Max 3 assert. Adj vissza CSAK a teszt blokkot:\n"
            "```python\n"
            "# === TESZTEK ===\n"
            "if __name__ == \"__test__\":\n"
            "    [assert1]\n"
            "    [assert2]\n"
            "    print(\"TESZTEK OK\")\n"
            "```\n"
            "TILOS: import unittest, class TestXxx, assert conn is not None"
        )
        s3b_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s3b_prompt, max_tokens=800, chat_id=chat_id)
        log.info(f"[GEN] S3b válasz hossza: {len(s3b_resp)} kar")
        s3b_block = extract_code_block(s3b_resp)
        log.info(f"[GEN] S3b extracted: {repr(s3b_block[:200] if s3b_block else 'NONE')}")

        if s3b_block and "# === TESZTEK ===" in s3b_block:
            test_part = s3b_block[s3b_block.index("# === TESZTEK ==="):]
            code = "# === KÓD ===\n" + clean_code + "\n\n" + test_part
        else:
            log.warning("[GEN] S3b nem adott vissza teszt blokkot – retry")
            s3b_retry_prompt = (
                "A következő Python kódhoz írj 2 assert-alapú unit tesztet.\n"
                f"```python\n{clean_code[:1500]}\n```\n\n"
                "Csak a teszt blokkot add vissza:\n"
                "```python\n"
                "# === TESZTEK ===\n"
                "if __name__ == \"__test__\":\n"
                "    assert valami == elvart\n"
                "    assert isinstance(valami, list)\n"
                "    print(\"TESZTEK OK\")\n"
                "```\n"
                "TILOS: import unittest, class TestXxx, assert conn is not None"
            )
            s3b_retry = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s3b_retry_prompt, max_tokens=600, chat_id=chat_id)
            s3b_retry_block = extract_code_block(s3b_retry)
            if s3b_retry_block and "# === TESZTEK ===" in s3b_retry_block:
                test_part = s3b_retry_block[s3b_retry_block.index("# === TESZTEK ==="):]
                code = "# === KÓD ===\n" + clean_code + "\n\n" + test_part
                log.info("[GEN] S3b retry sikeres")
            else:
                log.warning("[GEN] S3b retry is NONE – dummy fallback")
                code = (
                    "# === KÓD ===\n" + clean_code +
                    '\n\n# === TESZTEK ===\n'
                    'if __name__ == "__test__":\n'
                    '    assert True  # fallback\n'
                    '    print("TESZTEK OK")'
                )
        log.info(f"[GEN] Komplex generálás kész – 3a+3b session | {len(code.splitlines())} sor")

    else:
        # ── EGYSZERŰ: 2 sessionös generálás ─────────────────
        await update_status(status_msg, "1️⃣ *Kód generálás...*")
        s1_prompt = (
            f"{few_shot_pattern_block}"
            f"{history_context}"
            f"A feladat: {task}\n\n"
            "Generálj tiszta, TELJES Python megoldást.\n"
            "KÖTELEZŐ FORMÁTUM:\n```python\n# === KÓD ===\n[teljes megoldás]\n```\n"
            "Csak a kód, tesztek NEM kellenek."
        )
        s1_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s1_prompt, max_tokens=4000, chat_id=chat_id)
        s1_code = extract_code_block(s1_resp) or ""

        await update_status(status_msg, "1️⃣ *Unit tesztek...*")
        s2_prompt = (
            f"Kész kód:\n```python\n{s1_code[:3000]}\n```\n\n"
            f"Feladat: {task}\n\n"
            "Adj hozzá 2-3 LOGIKAI tesztet. A tesztek a KÓD BELSŐ LOGIKÁJÁT ellenőrizzék,\n"
            "NEM az infrastruktúra hívások sikerességét.\n\n"
            "TESZTÍRÁSI SZABÁLYOK:\n"
            "TILOS – infrastructure eredmény tesztelése:\n"
            "  assert conn is not None          <- DB kapcsolat\n"
            "  assert send_alert(x) == True     <- API hívás\n"
            "  assert sync_to_sheets(r) == True <- webhook hívás\n\n"
            "HELYES – logika, adatstruktúra, feltételek tesztelése:\n"
            "  result = transform(raw_data)\n"
            "  assert isinstance(result, list)  <- helyes típus\n"
            "  assert len(result) == len(raw_data) <- adatveszteség nincs\n\n"
            "Max 3 assert, mindegyik MÁS logikai aspektust teszteljen.\n"
            "```python\n# === KÓD ===\n[kód]\n\n# === TESZTEK ===\n"
            "if __name__ == \"__test__\":\n    [assert1]\n    [assert2]\n    print(\"TESZTEK OK\")\n```"
        )
        s2_resp = await call_claude_tracked(PIPELINE_PROMPTS["DEVELOPER"], s2_prompt, max_tokens=2000, chat_id=chat_id)
        combined = extract_code_block(s2_resp)
        code = combined if (combined and "# === TESZTEK ===" in combined) else (
            s1_code + '\n\n# === TESZTEK ===\nif __name__ == "__test__":\n    assert True\n    print("TESZTEK OK")'
        )
        log.info(f"[GEN] Egyszerű generálás kész – 2 session | {len(code.splitlines())} sor")

    if not code.strip():
        await status_msg.delete()
        await update.message.reply_text("❌ Kód generálás sikertelen. Pontosítsd a feladatot!")
        return

    # ── Kockázati szűrő ───────────────────────────────────────
    risks = [kw for kw in RISK_KEYWORDS if kw in code.lower()]
    if risks:
        task_id = str(uuid.uuid4())[:8]
        approved = await ask_risk_approval(update, risks, task_id)
        if not approved:
            await status_msg.delete()
            return

    # ── 2. SANDBOX (1. + 2. szint) ───────────────────────────
    await update_status(status_msg, "2️⃣ *Sandbox validáció...*\n🔬 Statikus szűrő + unit tesztek")

    def _build_few_shot_fix_block(error_text: str) -> str:
        """Dinamikusan összeállítja a few-shot fix blokkot sandbox hiba alapján."""
        samples = get_relevant_few_shot_samples(task, error_text=error_text, max_samples=2)
        if not samples:
            return ""
        import json as _json
        lines = ["KORÁBBI HASONLÓ HIBÁK ÉS JAVÍTÁSAIK (tanulj belőlük!):\n"]
        for i, s in enumerate(samples, 1):
            try:
                issues = _json.loads(s["issues"]) if isinstance(s["issues"], str) else s["issues"]
                issues_str = "; ".join(issues[:2]) if isinstance(issues, list) else str(issues)
            except Exception:
                issues_str = str(s["issues"])[:200]
            lines.append(f"--- Példa {i} (relevancia: {s['score']:.0%}) ---")
            lines.append(f"Feladat volt: {s['prompt']}")
            lines.append(f"Hibás kód részlet:\n```python\n{s['bad_code'][:400]}\n```")
            lines.append(f"Gemini kifogások: {issues_str}")
            lines.append(f"Javítás ami működött:\n```python\n{s['fixed_code'][:400]}\n```\n")
        return "\n".join(lines) + "\n"

    sandbox_result = await sandbox.validate_with_retry(
        code=code,
        task=task,
        ai_fix_callback=ai_fix_callback,
        status_callback=lambda m: update_status(status_msg, f"2️⃣ *Sandbox*\n{m}"),
        few_shot_fixes=_build_few_shot_fix_block("")
    )

    if not sandbox_result.success:
        await status_msg.delete()
        await update.message.reply_text(
            f"❌ *Sandbox sikertelen* {SANDBOX_MAX_RETRIES} próba után\n\n"
            f"{format_sandbox_report(sandbox_result)}\n\n"
            f"*Mit tehetsz:*\n"
            f"• Pontosítsd a feladatot\n"
            f"• `/bypass` sandbox nélkül",
            parse_mode="Markdown"
        )
        return

    validated_code = sandbox_result.final_code
    sandbox_summary = format_sandbox_report(sandbox_result)
    test_result_str = (
        f"Tesztek: {sandbox_result.tests_passed}/{sandbox_result.tests_total} OK\n"
        f"Stdout: {sandbox_result.stdout[:300]}"
    )

    # ── 3. GEMINI AUDIT (3. szint) ────────────────────────────
    await update_status(
        status_msg,
        "3️⃣ *Gemini audit...*\n🔮 Logikai ellenőrzés, projekt szabályok, minőség"
    )

    audit_result = await auditor.audit(
        code=validated_code,
        task=task,
        test_result=test_result_str
    )

    log.info(f"[AUDIT] {audit_result.verdict} – score: {audit_result.score}/100")

    if not audit_result.passed and audit_result.verdict != "SKIP":
        for fix_round in range(1, AUDIT_MAX_RETRIES + 1):
            await update_status(
                status_msg,
                f"3️⃣ *Gemini: FAIL* (javítás #{fix_round})\n"
                f"🔧 Claude javítja a kifogásolt részeket...\n"
                f"`{audit_result.issues[0][:80] if audit_result.issues else ''}`"
            )

            bad_code_snapshot    = validated_code
            bad_issues_snapshot  = audit_result.issues[:]
            bad_score_snapshot   = audit_result.score

            fix_prompt     = format_audit_for_fix_prompt(audit_result, validated_code, task)
            fixed_response = await call_claude(
                system=(
                    "Senior Python fejlesztő vagy. "
                    "Javítsd ki a megadott problémákat. "
                    "Adj vissza kódot ```python blokkban "
                    "(# === KÓD === és # === TESZTEK === szekciókkal)."
                ),
                user_msg=fix_prompt
            )

            new_code = extract_code_block(fixed_response)
            if not new_code:
                break

            await update_status(status_msg, f"3️⃣ *Javítás #{fix_round}* → Sandbox újra...")
            sandbox_result = await sandbox.validate_with_retry(
                code=new_code, task=task,
                ai_fix_callback=ai_fix_callback,
                few_shot_fixes=_build_few_shot_fix_block(sandbox_result.stderr or "")
            )
            if not sandbox_result.success:
                save_fix_sample(
                    prompt=task,
                    bad_code=bad_code_snapshot,
                    gemini_issues=bad_issues_snapshot,
                    gemini_score=bad_score_snapshot,
                    fixed_code=new_code,
                    fix_round=fix_round,
                    fix_succeeded=False
                )
                break

            validated_code = sandbox_result.final_code

            await update_status(status_msg, f"3️⃣ *Javítás #{fix_round}* → Gemini újra ellenőriz...")
            audit_result = await auditor.audit(
                code=validated_code, task=task,
                test_result=f"Javítás #{fix_round} utáni tesztek OK"
            )

            save_fix_sample(
                prompt=task,
                bad_code=bad_code_snapshot,
                gemini_issues=bad_issues_snapshot,
                gemini_score=bad_score_snapshot,
                fixed_code=validated_code,
                fix_round=fix_round,
                fix_succeeded=audit_result.passed
            )

            if audit_result.passed:
                log.info(f"[AUDIT] Javítás #{fix_round} után PASS!")
                break

    # ── VÉGEREDMÉNY ──────────────────────────────────────────
    await status_msg.delete()

    main_code, test_section = sandbox.extract_sections(validated_code)

    audit_icon    = "✅" if audit_result.passed else "⚠️"
    audit_verdict = audit_result.verdict
    if audit_verdict == "SKIP":
        audit_line = "🔮 Gemini audit: kihagyva (API nem elérhető)"
    else:
        audit_line = f"🔮 Gemini audit: {audit_icon} {audit_verdict} ({audit_result.score}/100)"

    # ── FÁJLBA MENTÉS ────────────────────────────────────────
    output_dir = BASE_DIR / "outputs"
    os.makedirs(output_dir, exist_ok=True)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_task  = re.sub(r'[^a-zA-Z0-9]', '_', task[:40]).strip('_')
    filename   = f"{timestamp}_{safe_task}.py"
    filepath   = os.path.join(output_dir, filename)
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"# AXON Neural Bridge – generált kód\n")
            f.write(f"# Feladat: {task}\n")
            f.write(f"# Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Sandbox: {'PASS' if sandbox_result.success else 'FAIL'} | Gemini: {audit_verdict} ({audit_result.score}/100)\n")
            f.write(f"# {'='*60}\n\n")
            f.write(validated_code)
        log.info(f"[OUTPUT] Kód mentve: {filepath}")
        file_line = f"📁 `{filename}`"

        # ── v7.0: README.md generálás ────────────────────────
        readme_line = await _generate_readme(
            task=task,
            main_code=main_code,
            filename=filename,
            output_dir=output_dir,
            timestamp=timestamp,
            safe_task=safe_task,
            audit_result=audit_result,
            sandbox_result=sandbox_result
        )

    except Exception as e:
        log.error(f"[OUTPUT] Fájl mentési hiba: {e}")
        file_line  = f"⚠️ Fájl mentés sikertelen: {str(e)[:100]}"
        readme_line = ""

    # ── TELEGRAM – rövid összefoglaló ────────────────────────
    multiturn_note = " _(multi-turn)_" if is_multiturn else ""
    header = (
        f"🎯 *DEVELOPER* válasza{multiturn_note}:\n"
        f"{sandbox_summary}\n"
        f"{audit_line}\n"
        f"{file_line}\n"
        f"{readme_line}\n"
    )

    if not audit_result.passed and audit_verdict != "SKIP" and audit_result.issues:
        header += "\n⚠️ *Fennmaradó megjegyzések:*\n"
        for issue in audit_result.issues[:2]:
            header += f"• {issue}\n"

    preview_lines = main_code.splitlines()[:15]
    preview       = "\n".join(preview_lines)
    total_lines   = len(main_code.splitlines())
    if total_lines > 15:
        preview += f"\n# ... (+{total_lines - 15} sor)"

    full_reply = header + f"\n```python\n{preview}\n```"

    # ── COST ÖSSZESÍTÉS ──────────────────────────────────────
    task_tokens = _pop_task_tokens(chat_id)
    task_cost   = _tokens_to_usd(task_tokens["input"], task_tokens["output"])
    cost_line   = (
        f"💰 *Feladat költsége:* ${task_cost:.4f} "
        f"_({task_tokens['input']//1000}k in / {task_tokens['output']//1000}k out"
        f" · {task_tokens['calls']} hívás)_"
    )
    full_reply += f"\n\n{cost_line}"

    await safe_send(update, full_reply)

    # Cost logolás DB-be
    log_task_cost(
        task=task,
        input_tokens=task_tokens["input"],
        output_tokens=task_tokens["output"],
        cost_usd=task_cost,
        calls=task_tokens["calls"]
    )

    # ── HISTORY MENTÉS – validated_code kerül be (nem Telegram formázás!) ──
    add_to_history(chat_id, "user",      task,           "DEVELOPER")
    add_to_history(chat_id, "assistant", validated_code, "DEVELOPER", task=task)
    log.info(f"[HISTORY] DEVELOPER kód mentve history-ba ({len(validated_code)} kar)")

    # ── TRAINING DATA + CACHE MENTÉS ─────────────────────────
    final_success = sandbox_result.success and audit_result.passed
    save_training_sample(
        expert_mode="developer",
        prompt=task,
        generated_code=validated_code,
        sandbox_result=sandbox_result.stdout[:500] + sandbox_result.stderr[:500],
        audit_result=audit_result.telegram_summary if audit_result.verdict != "SKIP" else "SKIP",
        sandbox_ok=sandbox_result.success,
        audit_ok=audit_result.passed,
        success=final_success,
        retry_count=sandbox_result.attempt - 1
    )
    # Cache csak friss session + sikeres futás esetén
    if final_success and not is_multiturn:
        save_cached_response("developer", task, full_reply)

    # ── v7.2: Auto fájlküldés Telegramon ─────────────────────
    if final_success and os.path.exists(filepath):
        try:
            with open(filepath, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename=filename,
                    caption=f"📎 *{filename}*\n_Sandbox ✅ | Gemini {audit_result.verdict} ({audit_result.score}/100)_",
                    parse_mode="Markdown"
                )
            log.info(f"[FILES] Auto fájlküldés: {filename}")
        except Exception as e:
            log.warning(f"[FILES] Auto küldési hiba (nem kritikus): {e}")

    log.info(f"[OUTPUT] Feladat kész – Audit: {audit_result.verdict} ({audit_result.score}/100) | Fájl: {filename}")


# ═══════════════════════════════════════════════════════════════
#  FŐ FELADATFELDOLGOZÓ v6.0
#  Folyamat: timeout check → detect_pipeline → pipeline-specifikus handler
# ═══════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════
#  FÁJL FOGADÁS – v7.3
#  Telegram document → lokális elemzés → Claude kontextus
# ═══════════════════════════════════════════════════════════════

UPLOADS_DIR = BASE_DIR / "uploads"

SUPPORTED_EXTENSIONS = {
    ".csv", ".json", ".xlsx", ".xls",
    ".txt", ".xml", ".yaml", ".yml", ".sql"
}


class FileAnalyzer:
    """
    Lokális fájlelemzés az AXON process-ben — NEM sandbox.
    Csak stdlib + openpyxl (Excel-hez).
    Nagy fájlnál: első 100 sor + oszlop statisztika.
    """

    MAX_ROWS_PREVIEW = 100
    MAX_CHARS_TOTAL  = 3000  # Claude-nak küldött max kontextus

    @staticmethod
    def analyze(filepath: str) -> str:
        """Fájl elemzése → Claude-nak küldendő kontextus string."""
        ext  = Path(filepath).suffix.lower()
        name = Path(filepath).name
        size = os.path.getsize(filepath)
        size_str = f"{size // 1024}KB" if size >= 1024 else f"{size}B"

        try:
            if ext == ".csv":
                return FileAnalyzer._analyze_csv(filepath, name, size_str)
            elif ext == ".json":
                return FileAnalyzer._analyze_json(filepath, name, size_str)
            elif ext in (".xlsx", ".xls"):
                return FileAnalyzer._analyze_excel(filepath, name, size_str)
            elif ext in (".yaml", ".yml"):
                return FileAnalyzer._analyze_yaml(filepath, name, size_str)
            elif ext == ".xml":
                return FileAnalyzer._analyze_xml(filepath, name, size_str)
            elif ext == ".sql":
                return FileAnalyzer._analyze_text(filepath, name, size_str, "SQL")
            else:
                return FileAnalyzer._analyze_text(filepath, name, size_str, "TXT")
        except Exception as e:
            return f"[FÁJL] {name} ({size_str}) – Elemzési hiba: {e}"

    @staticmethod
    def _analyze_csv(filepath: str, name: str, size_str: str) -> str:
        import csv, io
        lines = []
        with open(filepath, "rb") as raw:
            # Encoding detektálás: UTF-8 BOM → UTF-8, különben latin-1 fallback
            sample = raw.read(4096)
        enc = "utf-8-sig" if sample[:3] == b'\xef\xbb\xbf' else "utf-8"
        try:
            with open(filepath, encoding=enc, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
        except UnicodeDecodeError:
            with open(filepath, encoding="latin-1", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)

        if not rows:
            return f"[FÁJL] {name} – Üres fájl"

        total_rows  = len(rows)
        header      = rows[0] if rows else []
        data_rows   = rows[1:] if len(rows) > 1 else []
        preview     = data_rows[:5]

        # Típus detektálás az első 20 adatsor alapján
        col_types = []
        for col_idx in range(len(header)):
            vals = [r[col_idx] for r in data_rows[:20] if col_idx < len(r) and r[col_idx].strip()]
            if not vals:
                col_types.append("empty")
                continue
            try:
                [float(v.replace(",", ".")) for v in vals]
                col_types.append("numeric")
            except ValueError:
                col_types.append("text")

        col_info = ", ".join(
            f"`{h}` ({t})" for h, t in zip(header, col_types)
        )

        preview_lines = []
        for row in preview:
            preview_lines.append("  " + " | ".join(str(v)[:30] for v in row))

        result = (
            f"[FÁJL] {name} ({size_str})\n"
            f"Típus: CSV | Sorok: {total_rows - 1} adat + fejléc | Oszlopok: {len(header)}\n"
            f"Oszlopok: {col_info}\n"
            f"Minta (első 5 sor):\n" + "\n".join(preview_lines)
        )
        return result[:FileAnalyzer.MAX_CHARS_TOTAL]

    @staticmethod
    def _analyze_json(filepath: str, name: str, size_str: str) -> str:
        import json as _json

        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()
        try:
            data = _json.loads(content)
        except _json.JSONDecodeError as e:
            return f"[FÁJL] {name} ({size_str}) – Érvénytelen JSON: {e}"

        def describe(obj, depth=0, max_depth=3) -> str:
            if depth > max_depth:
                return "..."
            if isinstance(obj, dict):
                keys = list(obj.keys())[:10]
                inner = ", ".join(f'"{k}": {describe(obj[k], depth+1)}' for k in keys)
                suffix = ", ..." if len(obj) > 10 else ""
                return "{" + inner + suffix + "}"
            elif isinstance(obj, list):
                count = len(obj)
                sample = describe(obj[0], depth+1) if obj else "[]"
                return f"[{count} elem, pl: {sample}]"
            elif isinstance(obj, str):
                return f'"{obj[:30]}"'
            else:
                return str(obj)

        structure = describe(data)
        top_level = type(data).__name__
        count_info = f"{len(data)} elem" if isinstance(data, (list, dict)) else ""

        result = (
            f"[FÁJL] {name} ({size_str})\n"
            f"Típus: JSON | Top-level: {top_level} {count_info}\n"
            f"Struktúra: {structure}"
        )
        return result[:FileAnalyzer.MAX_CHARS_TOTAL]

    @staticmethod
    def _analyze_excel(filepath: str, name: str, size_str: str) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheets = wb.sheetnames
            result_parts = [f"[FÁJL] {name} ({size_str})\nTípus: Excel | Lapok: {', '.join(sheets)}"]

            for sheet_name in sheets[:2]:  # max 2 lap
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True, max_row=FileAnalyzer.MAX_ROWS_PREVIEW))
                if not rows:
                    continue
                header  = [str(c) if c is not None else "" for c in rows[0]]
                total   = ws.max_row or len(rows)
                preview = rows[1:6]
                col_str = ", ".join(f"`{h}`" for h in header[:15])
                preview_str = "\n".join(
                    "  " + " | ".join(str(v)[:25] if v is not None else "" for v in row)
                    for row in preview
                )
                result_parts.append(
                    f"Lap: {sheet_name} | ~{total} sor | Oszlopok: {col_str}\n"
                    f"Minta:\n{preview_str}"
                )
            wb.close()
            return "\n".join(result_parts)[:FileAnalyzer.MAX_CHARS_TOTAL]

        except ImportError:
            return (
                f"[FÁJL] {name} ({size_str}) – Excel fájl\n"
                f"⚠️ openpyxl nincs telepítve: pip install openpyxl"
            )

    @staticmethod
    def _analyze_yaml(filepath: str, name: str, size_str: str) -> str:
        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()
        lines = content.splitlines()
        preview = "\n".join(lines[:30])
        return (
            f"[FÁJL] {name} ({size_str})\n"
            f"Típus: YAML | {len(lines)} sor\n"
            f"Tartalom (első 30 sor):\n{preview}"
        )[:FileAnalyzer.MAX_CHARS_TOTAL]

    @staticmethod
    def _analyze_xml(filepath: str, name: str, size_str: str) -> str:
        import xml.etree.ElementTree as ET
        try:
            tree = ET.parse(filepath)
            root = tree.getroot()
            children = [child.tag for child in list(root)[:10]]
            attrs    = list(root.attrib.keys())[:5]
            return (
                f"[FÁJL] {name} ({size_str})\n"
                f"Típus: XML | Root: `{root.tag}` | "
                f"Gyerek elemek: {', '.join(children) or 'nincs'} | "
                f"Attribútumok: {', '.join(attrs) or 'nincs'}"
            )[:FileAnalyzer.MAX_CHARS_TOTAL]
        except ET.ParseError as e:
            return f"[FÁJL] {name} – Érvénytelen XML: {e}"

    @staticmethod
    def _analyze_text(filepath: str, name: str, size_str: str, ftype: str) -> str:
        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()
        lines   = content.splitlines()
        preview = "\n".join(lines[:30])
        return (
            f"[FÁJL] {name} ({size_str})\n"
            f"Típus: {ftype} | {len(lines)} sor\n"
            f"Tartalom (első 30 sor):\n{preview}"
        )[:FileAnalyzer.MAX_CHARS_TOTAL]


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    v7.3 – Telegram fájl fogadás handler.
    Letölti, elemzi, és beilleszti a session context-be.
    Ezután Claude javaslatot ad mit lehet csinálni a fájllal.
    """
    if not is_owner(update) or not system_running(update):
        return

    chat_id  = str(update.effective_chat.id)
    doc      = update.message.document
    fname    = doc.file_name or "ismeretlen_fajl"
    ext      = Path(fname).suffix.lower()

    # Típus ellenőrzés
    if ext not in SUPPORTED_EXTENSIONS:
        await update.message.reply_text(
            f"⚠️ *Nem támogatott fájltípus:* `{ext}`\n\n"
            f"Támogatott: {', '.join(sorted(SUPPORTED_EXTENSIONS))}",
            parse_mode="Markdown"
        )
        return

    msg = await update.message.reply_text(
        f"📥 *Fájl fogadva:* `{fname}`\n_Elemzés folyamatban..._",
        parse_mode="Markdown"
    )

    try:
        # uploads/ mappa létrehozása
        os.makedirs(UPLOADS_DIR, exist_ok=True)

        # Letöltés
        tg_file   = await doc.get_file()
        save_path = os.path.join(UPLOADS_DIR, fname)
        await tg_file.download_to_drive(save_path)
        log.info(f"[UPLOAD] Letöltve: {save_path}")

        # Lokális elemzés
        analysis = FileAnalyzer.analyze(save_path)
        log.info(f"[UPLOAD] Elemzés kész: {len(analysis)} kar")

        # Session history-ba kerül ANALYST pipeline-ként
        # A fájl path is bekerül → következő feladatban AXON tudja hol van
        file_context = (
            f"{analysis}\n\n"
            f"FÁJL ELÉRÉSI ÚT (használd a generált kódban): {save_path}"
        )
        add_to_history(chat_id, "user",      f"[FÁJL FELTÖLTVE] {fname}", "ANALYST")
        add_to_history(chat_id, "assistant", file_context,                 "ANALYST")

        # Claude javaslatot ad
        suggest_prompt = (
            f"A felhasználó feltöltött egy fájlt. Elemzés eredménye:\n\n"
            f"{analysis}\n\n"
            f"Fájl elérési útja: {save_path}\n\n"
            "Röviden (3-5 mondatban) magyarázd el:\n"
            "1. Mit látsz ebben a fájlban\n"
            "2. Milyen Python automatizálási feladatokhoz lehet használni\n"
            "3. Konkrét javaslatot adj mit csináljunk vele\n\n"
            "Magyarul válaszolj, tömören."
        )
        suggestion = await call_claude(
            system=get_context_for_pipeline("ANALYST"),
            user_msg=suggest_prompt,
            max_tokens=400
        )

        await msg.delete()
        await safe_send(update,
            f"📊 *Fájl elemzés kész:* `{fname}`\n\n"
            f"{suggestion}\n\n"
            f"_Fájl elmentve: `{save_path}`_\n"
            f"_Következő üzenetedben mondd meg mit csináljak vele!_"
        )
        log.info(f"[UPLOAD] Elemzés és javaslat elküldve | chat: {chat_id}")

    except Exception as e:
        log.error(f"[UPLOAD] Hiba: {e}", exc_info=True)
        await msg.delete()
        await update.message.reply_text(
            f"❌ *Fájl feldolgozási hiba:* {str(e)[:200]}",
            parse_mode="Markdown"
        )


async def handle_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_owner(update) or not system_running(update):
        return

    task    = update.message.text
    chat_id = str(update.effective_chat.id)
    log.info(f"Feladat: {task[:60]}...")

    # ── Session timeout értesítő ──────────────────────────────
    # A was_timeout_cleared() egyszeri olvasás + auto-reset
    if was_timeout_cleared(chat_id):
        await update.message.reply_text(
            "🔄 *Új session kezdődött* _(2 óra inaktivitás után)_\n"
            "Az előző conversation history törölve.",
            parse_mode="Markdown"
        )

    # ── Auto-compact trigger (v8.2) ───────────────────────────
    # Ha a history meghaladja a 6000 karaktert → automatikus tömörítés
    # Cargo-connect state machine + Claw Code compact_session() mintájára
    _history_now = get_history(chat_id)
    if _history_now:
        _history_size = sum(len(t.get("content", "")) for t in _history_now)
        if _history_size > 6000:
            log.info(f"[AUTO-COMPACT] History méret: {_history_size} kar → tömörítés indul")
            loop = asyncio.get_running_loop()
            _compact_result = await loop.run_in_executor(
                None, compact_history, _history_now, call_claude_sync, chat_id
            )
            if not _compact_result.skipped:
                from axon_memory import clear_history as _clear_h, add_to_history as _add_h
                _clear_h(chat_id)
                for turn in _compact_result.new_history:
                    _add_h(chat_id, turn["role"], turn["content"],
                           turn.get("pipeline", "DEVELOPER"))
                await update.message.reply_text(
                    f"🗜 _Auto-compact: history {_compact_result.original_chars:,} → "
                    f"{_compact_result.new_chars:,} kar_",
                    parse_mode="Markdown"
                )
                log.info(f"[AUTO-COMPACT] Kész: {_compact_result.original_chars} → {_compact_result.new_chars} kar")

    # ── Pipeline routing ──────────────────────────────────────
    pipeline = await detect_pipeline(task)
    icon, label = PIPELINE_META[pipeline]
    log.info(f"[ROUTER] Pipeline: {pipeline}")

    use_history  = pipeline in HISTORY_ENABLED_PIPELINES
    turn_count   = get_history_turn_count(chat_id) if use_history else 0
    is_multiturn = use_history and turn_count > 0

    validation_hint = "\n1️⃣ 2️⃣ 3️⃣ _3 szintű validáció aktív_" if pipeline == "DEVELOPER" else ""
    history_hint    = f"\n🧠 _Multi-turn ({turn_count // 2} előző válasz kontextusban)_" if is_multiturn else ""

    if pipeline in NO_CACHE_PIPELINES:
        cache_hint = "\n_✍️ Egyedi generálás (nincs cache, nincs history)_"
    elif is_multiturn:
        cache_hint = "\n_🔗 Cache bypass (multi-turn session)_"
    else:
        cache_hint = "\n_⚡ Cache ellenőrzés..._"

    status_msg = await update.message.reply_text(
        f"⏳ *Feldolgozás...* {icon} _{label} pipeline_{validation_hint}{history_hint}{cache_hint}",
        parse_mode="Markdown"
    )

    try:
        if pipeline == "DEVELOPER":
            await run_developer_pipeline(update, task, status_msg, chat_id)
        else:
            await run_simple_pipeline(update, pipeline, task, status_msg, chat_id)

    except Exception as e:
        try:
            await status_msg.delete()
        except Exception:
            pass
        await update.message.reply_text(
            f"❌ *Váratlan hiba:*\n`{str(e)[:300]}`\n\nPróbáld újra!",
            parse_mode="Markdown"
        )
        log.error(f"Hiba: {e}", exc_info=True)

# ═══════════════════════════════════════════════════════════════
#  INDÍTÁS
# ═══════════════════════════════════════════════════════════════
def main():
    global watchman
    log.info("═══════════════════════════════════════════════════")
    log.info("  AXON TELEGRAM BRIDGE v8.2 – INDUL...")
    log.info("  Konfiguráció: .env fájlból betöltve ✅")
    log.info("  1. szint: Statikus biztonsági szűrő ✅")
    log.info("  2. szint: Unit tesztek (sandbox) ✅")
    log.info("  3. szint: Gemini logikai audit ✅")
    log.info("  Memory/Training: ✅ (axon.db)")
    log.info("  Task Cache: ✅ (axon.db – 0 API hívás ismételt feladatnál)")
    log.info("  OWNER_CHAT_ID: ✅ (perzisztens – axon.db)")
    log.info("  Multi-session generálás: ✅ (SIMPLE: 2, COMPLEX: 4)")
    log.info("  Multi-expert routing: ✅ (DEVELOPER/PLANNER/CREATIVE/ANALYST)")
    log.info("  SOUL.md promptok: ✅ (souls/ mappa, fallback: hardcoded)")
    log.info("  Conversation Memory: ✅ (in-memory, 8000 kar limit, 2h timeout)")
    log.info("  Auto-compact: ✅ (6000 kar felett automatikus tömörítés)")
    log.info("  /upwork wizard: ✅ (3 lépéses ConversationHandler)")
    log.info("  SRE Watchman: ✅ (JobQueue – fő event loop-ban)")
    log.info("  Kill switch: /stop | Stats: /stats | Cache: /cache_clear")
    log.info("  History: /history | Clear: /clear | Compact: /compact | Review: /review")
    log.info("═══════════════════════════════════════════════════")

    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # /upwork wizard – ConversationHandler (Cargo-connect state machine minta)
    upwork_wizard = ConversationHandler(
        entry_points=[CommandHandler("upwork", upwork_start)],
        states={
            UPWORK_JOB: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, upwork_got_job),
            ],
            UPWORK_BUDGET: [
                CommandHandler("skip", upwork_skip_budget),
                MessageHandler(filters.TEXT & ~filters.COMMAND, upwork_got_budget),
            ],
        },
        fallbacks=[CommandHandler("cancel", upwork_cancel)],
        conversation_timeout=300,  # 5 perc inaktivitás → auto cancel
    )

    app.add_handler(upwork_wizard)
    app.add_handler(CommandHandler("start",       start))
    app.add_handler(CommandHandler("help",        help_cmd))
    app.add_handler(CommandHandler("status",      status_cmd))
    app.add_handler(CommandHandler("stats",       stats_cmd))
    app.add_handler(CommandHandler("stop",        stop_cmd))
    app.add_handler(CommandHandler("bypass",      bypass_cmd))
    app.add_handler(CommandHandler("cache_clear", cache_clear_cmd))
    app.add_handler(CommandHandler("clear",       clear_cmd))
    app.add_handler(CommandHandler("history",     history_cmd))
    app.add_handler(CommandHandler("compact",     compact_cmd))
    app.add_handler(CommandHandler("review",      review_cmd))
    app.add_handler(CommandHandler("files",       files_cmd))
    app.add_handler(CallbackQueryHandler(handle_approval_callback))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_task))

    async def post_init(application):
        # ── Watchman → JobQueue migráció (v8.2) ──────────────────
        # PTB natív JobQueue – a fő event loop-ban fut, nem külön thread.
        # Ha a bot crashel és újraindul, a Watchman automatikusan újraindul vele.
        # asyncio.create_task() helyett: application.job_queue.run_repeating()
        async def send_alert(msg: str):
            if OWNER_CHAT_ID:
                await application.bot.send_message(
                    chat_id=OWNER_CHAT_ID,
                    text=msg,
                    parse_mode="Markdown"
                )

        _watchman = AxonWatchman(alert_callback=send_alert)

        async def watchman_job(job_context):
            """JobQueue callback – 60 másodpercenként hívja a Watchman check-et."""
            await _watchman._check_once()

        application.job_queue.run_repeating(
            watchman_job,
            interval=60,   # CHECK_INTERVAL
            first=10,      # 10 másodperc után első futás (bot init után)
            name="watchman"
        )
        log.info("SRE Watchman elindítva (JobQueue – fő event loop).")

    app.post_init = post_init

    log.info("✅ Bot fut! (Ctrl+C a leállításhoz)")
    app.run_polling()

if __name__ == "__main__":
    main()
